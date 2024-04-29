from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes
from telegram.ext import MessageHandler, filters
from telegram.ext import CallbackContext
import telegram
from telegram import Bot

import httpx
from functools import partial
import datetime
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import asyncio
import os

token = "7099765163:AAFSdTNCqJgwDgIral48IfZO9fJQLBj8Zx0"
client = httpx.AsyncClient(timeout=30.0) # Установка таймаута в 30 секунд
bot = Bot(token=token)

def save_schedule_time_for_user(chat_id, time):
    filename = 'users.xlsx'
    sheetname = 'Schedule'
    # Создаем новую книгу или загружаем существующую
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname
        # Заголовки для столбцов
        ws['A1'] = 'Chat ID'
        ws['B1'] = 'Scheduled Time'
        wb.save(filename)
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheetname]
    # Проверяем, существует ли уже такой id чата
    chat_id_exists = False
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        if row[0].value == chat_id:
            chat_id_exists = True
            break
    # Если id чата существует, обновляем время
    if chat_id_exists:
        for row in ws.iter_rows(min_row=2):
            if row[0].value == chat_id:
                row[1].value = time
                break
    else:
        # Если id чата не существует, добавляем новую строку
        ws.append([chat_id, time])
    # Сохраняем изменения
    wb.save(filename)

# Функция для определения четности недели
def is_even_week():
    return datetime.now().isocalendar()[1] % 2 == 1


def get_cabinet_for_subject(subject):
    file_name = "cabinets.xlsx"
    # Загрузка расписания из файла
    schedule_df = pd.read_excel(file_name, engine='openpyxl')
    # Проверяем, есть ли в DataFrame столбец с названием 'предмет'
    if 'предмет' not in schedule_df.columns:
        return "Столбец 'предмет' не найден в DataFrame."
    # Ищем строку с совпадающим предметом
    matching_row = schedule_df[schedule_df['предмет'] == subject]
    # Если совпадение найдено, извлекаем данные из столбца 'кабинет'
    if not matching_row.empty:
        cabinets = matching_row.iloc[0]['кабинет']
        return cabinets
    else:
        return "Предмет не найден в расписании."
# Функция для получения расписания на день
def get_schedule_for_day(day):
    # Определение чётности недели
    week_num = datetime.now().isocalendar()[1]
    # Выбор соответствующего файла на основе четности недели
    file_name = "S26Fchetn.xlsx" if week_num % 2 != 0 else "S26Fnech.xlsx"
    # Загрузка расписания из файла
    schedule_df = pd.read_excel(file_name, engine='openpyxl')
    # Проверяем, есть ли в DataFrame столбец с названием дня недели
    if day not in schedule_df.columns:
        return "Расписание для этого дня не найдено."
    # Получаем расписание для этого дня
    schedule_str = ''
    for index, row in schedule_df.iterrows():
        time_slot = row['Время']
        subject = row[day] if pd.notnull(row[day]) else ''
        # Если в ячейке есть предмет, добавляем его в расписание
        if subject:
            schedule_str += f"<u>{subject}</u>\n <i>┗ {time_slot} {get_cabinet_for_subject(subject)}</i>\n"
    return schedule_str.strip()  # Удаление лишних пробелов и переносов строк

# Функция для получения расписания на всю неделю
def get_schedule_for_week(is_current_week=True):
    # Определение чётности недели
    week_num = datetime.now().isocalendar()[1]
    if not is_current_week:
        week_num += 1
    # Выбор соответствующего файла на основе четности недели
    file_name = "S26Fchetn.xlsx" if week_num % 2 != 0 else "S26Fnech.xlsx"
    # Загрузка расписания из файла
    schedule_df = pd.read_excel(file_name, engine='openpyxl')
    schedule_str = ''
    # Список дней недели
    days_of_week = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
    # Перебор дней недели и соответствующих колонок в файле Excel
    for day in days_of_week:
        if day in schedule_df.columns:
            schedule_str += f"<u><b>❢{day}❢</b></u>\n"  # Заголовок для дня недели
            # Получение временных слотов и пар для дня недели
            for index, row in schedule_df.iterrows():
                time_slot = row['Время']
                subject = row[day] if pd.notnull(row[day]) else ''
                # Если в ячейке есть предмет, добавляем его в расписание
                if subject:
                    schedule_str += f"{subject}\n  <i>┗ {time_slot} {get_cabinet_for_subject(subject)}</i>\n"
            schedule_str += "\n"  # Добавляем разделитель между днями
    return schedule_str.strip()  # Удаление лишних пробелов и переносов строк

reminders = {}
user_states = {} # Добавляем словарь для отслеживания состояния пользователя

async def set_reminder_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = update.message.chat_id
    reminders[chat_id] = {'reminder_name': update.message.text}
    await update.message.reply_text('Введите дату и время, когда вы хотите получить напоминание в формате ГГГГ-ММ-ДД чч:мм:сс.')

async def send_reminder(chat_id, reminder_name, delay):
    await asyncio.sleep(delay)
    # Используйте экземпляр Bot для отправки сообщения
    await bot.send_message(chat_id, f'Время получить ваше напоминание "{reminder_name}"!')

async def subscribe(flag, update: Update, context: CallbackContext):
    while flag:
        now = datetime.now()
        if now.hour == 19 and now.minute == 0:
            await send_schedule(update, context)
        await asyncio.sleep(60)

async def send_schedule(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    # Вычисляем, какой завтра день недели
    tomorrow = datetime.now() + timedelta(days=1)
    days_of_week = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
    next_day_of_week = days_of_week[tomorrow.weekday()]
    # Получаем расписание для этого дня
    schedule = get_schedule_for_day(next_day_of_week)
    await update.message.reply_text(schedule, parse_mode='HTML')

async def get_schedule_for_days(chat_id, date_str, delay):
    await asyncio.sleep(delay)
    # Преобразование строки даты в объект datetime
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
    # Получение дня недели (0 - понедельник, 6 - воскресенье)
    day_of_week = date_obj.weekday()
    # Получение названия дня недели по номеру дня
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    day_name = days[day_of_week]
    # Определение чётности недели для заданной даты
    week_num = date_obj.isocalendar()[1]
    file_name = "S26Fchetn.xlsx" if week_num % 2 != 0 else "S26Fnech.xlsx"
    # Загрузка расписания из файла
    schedule_df = pd.read_excel(file_name, engine='openpyxl')
    # Проверяем, есть ли в DataFrame столбец с названием дня недели
    if day_name not in schedule_df.columns:
        await bot.send_message(chat_id, "Расписание для этого дня не найдено.")
    # Получаем расписание для этого дня
    schedule_str = f"Напоминаю! Отработка по этому расписанию:\n"
    for index, row in schedule_df.iterrows():
        time_slot = row['Время']
        subject = row[day_name] if pd.notnull(row[day_name]) else ''
        # Если в ячейке есть предмет, добавляем его в расписание
        if subject:
            schedule_str += f"<u>{subject}</u>\n <i>┗ {time_slot} {get_cabinet_for_subject(subject)}</i>\n"
    await bot.send_message(chat_id, text=schedule_str.strip(),  parse_mode='HTML')  # Удаление лишних пробелов и переносов строк

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    keyboard = [['Сегодня', 'Завтра'], ['Текущая неделя', 'Следующая неделя'],['Напоминание', 'Отработка'], ['Авто-рассылка', 'Отмена рассылки']]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text('Выберите опцию:', reply_markup=reply_markup)

# Асинхронная функция для обработки текстовых сообщений
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = update.message.text.lower()
    today = datetime.now()
    days_of_week = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

    if text == 'сегодня':
        current_day_of_week = days_of_week[today.weekday()]
        schedule = get_schedule_for_day(current_day_of_week)
        await update.message.reply_text(schedule, parse_mode='HTML')
    elif text == 'завтра':
        next_day = today + timedelta(days=1)
        next_day_of_week = days_of_week[next_day.weekday()]
        schedule = get_schedule_for_day(next_day_of_week)
        await update.message.reply_text(schedule, parse_mode='HTML')
    elif text == 'текущая неделя':
        schedule = get_schedule_for_week(is_current_week=True)
        await update.message.reply_text(schedule, parse_mode='HTML')
    elif text == 'следующая неделя':
        schedule = get_schedule_for_week(is_current_week=False)
        await update.message.reply_text(schedule, parse_mode='HTML')
    chat_id = update.message.chat_id
    text = update.message.text.lower()

    if text == 'напоминание':
        user_states[chat_id] = 'waiting_for_reminder_name'
        await update.message.reply_text('Введите название напоминания:')
    elif user_states.get(chat_id) == 'waiting_for_reminder_name':
        user_states[chat_id] = 'waiting_for_reminder_time'
        reminders[chat_id] = {'reminder_name': text}
        await update.message.reply_text(
            'Введите дату и время, когда вы хотите получить напоминание в формате ГГГГ-ММ-ДД чч:мм:сс.')
    elif user_states.get(chat_id) == 'waiting_for_reminder_time':
        try:
            reminder_time = datetime.strptime(text, '%Y-%m-%d %H:%M:%S')
            now = datetime.now()
            delta = reminder_time - now
            if delta.total_seconds() <= 0:
                await update.message.reply_text('Вы ввели прошедшую дату, попробуйте еще раз.')
            else:
                reminder_name = reminders[chat_id]['reminder_name']
                await update.message.reply_text(f'Напоминание "{reminder_name}" установлено на {reminder_time}.')
                asyncio.create_task(send_reminder(chat_id, reminder_name, delta.total_seconds()))
                user_states[chat_id] = None  # Сбрасываем состояние пользователя после установки напоминания
        except ValueError:
            await update.message.reply_text('Вы ввели неверный формат даты и времени, попробуйте еще раз.')

    if text == 'отработка':
        user_states[chat_id] = 'waiting_for_day_name'
        await update.message.reply_text('Введите день, за который будет отработка, в формате ГГГГ-ММ-ДД:')
    elif user_states.get(chat_id) == 'waiting_for_day_name':
        user_states[chat_id] = 'waiting_for_work_time'
        reminders[chat_id] = {'day_name': text}
        await update.message.reply_text(
            'Введите дату и время, когда вы хотите получить напоминание об отработке в формате ГГГГ-ММ-ДД чч:мм:сс.')
    elif user_states.get(chat_id) == 'waiting_for_work_time':
        try:
            work_time = datetime.strptime(text, '%Y-%m-%d %H:%M:%S')
            now = datetime.now()
            delta = work_time - now
            if delta.total_seconds() <= 0:
                await update.message.reply_text('Вы ввели прошедшую дату, попробуйте еще раз.')
            else:
                day_name = reminders[chat_id]['day_name']
                await update.message.reply_text(f'Напоминание об отработке за "{day_name}" установлено на {work_time}.')
                asyncio.create_task(get_schedule_for_days(chat_id, day_name, delta.total_seconds()))
                user_states[chat_id] = None  # Сбрасываем состояние пользователя после установки напоминания
        except ValueError:
            await update.message.reply_text('Вы ввели неверный формат даты и времени, попробуйте еще раз.')

    if text == 'авто-рассылка':
        await subscribe(True, update, context)
    if text == 'отмена рассылки':
        await subscribe(False, update, context)

def main():
    # Создаем объект Application и передаем ему токен вашего бота.
    application = Application.builder().token(token).build()

    # Регистрируем обработчики команд и текстовых сообщений
    application.add_handler(CommandHandler('start', start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))


    # Запускаем бота
    application.run_polling()

if __name__ == '__main__':
    main()
